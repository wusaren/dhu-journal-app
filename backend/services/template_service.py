"""
模板服务 - 处理Excel模板文件的上传、表头识别和模糊匹配
"""
import os
import logging
from typing import List, Dict, Optional, Tuple
from pathlib import Path
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# 系统字段映射定义
SYSTEM_FIELD_MAPPINGS = {
    'manuscript_id': {
        'label': '稿件号',
        'category': '基本信息',
        'keywords': ['稿件号', '稿件编号', '稿件', 'manuscript', 'id']
    },
    'pdf_pages': {
        'label': '页数',
        'category': '基本信息',
        'keywords': ['页数', '页', 'pages', '总页数']
    },
    'first_author': {
        'label': '一作',
        'category': '作者信息',
        'keywords': ['一作', '第一作者', 'first author', 'first']
    },
    'corresponding': {
        'label': '通讯',
        'category': '作者信息',
        'keywords': ['通讯', '通讯作者', 'corresponding', 'correspondence']
    },
    'authors': {
        'label': '作者',
        'category': '作者信息',
        'keywords': ['作者', 'author', 'authors']
    },
    'issue': {
        'label': '刊期',
        'category': '基本信息',
        'keywords': ['刊期', '期', 'issue']
    },
    'is_dhu': {
        'label': '是否东华大学',
        'category': '基本信息',
        'keywords': ['是否东华', '东华', 'dhu', 'is_dhu']
    },
    'title': {
        'label': '标题',
        'category': '论文信息',
        'keywords': ['标题', '题目', 'title']
    },
    'chinese_title': {
        'label': '中文标题',
        'category': '论文信息',
        'keywords': ['中文标题', '中文题目', 'chinese title']
    },
    'chinese_authors': {
        'label': '中文作者',
        'category': '作者信息',
        'keywords': ['中文作者', 'chinese author', 'chinese authors']
    },
    'doi': {
        'label': 'DOI',
        'category': '论文信息',
        'keywords': ['doi', 'DOI']
    },
    'page_start': {
        'label': '起始页码',
        'category': '基本信息',
        'keywords': ['起始页', '起始页码', '开始页', 'page start', 'start page']
    },
    'page_end': {
        'label': '结束页码',
        'category': '基本信息',
        'keywords': ['结束页', '结束页码', '终止页', 'page end', 'end page']
    },
}

class TemplateService:
    """模板服务类"""
    
    def extract_headers_from_excel(self, file_path: str) -> List[str]:
        """
        从Excel文件中提取表头（自动跳过空行，查找第一个有内容的行）
        """
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            
            max_search_rows = min(10, ws.max_row)
            header_row = None
            
            for row_idx in range(1, max_search_rows + 1):
                row = ws[row_idx]
                has_content = False
                for cell in row:
                    if cell.value and str(cell.value).strip():
                        has_content = True
                        break
                
                if has_content:
                    header_row = row_idx
                    break
            
            if header_row is None:
                logger.warning(f"Excel文件前{max_search_rows}行都没有内容，无法提取表头")
                return []
            
            headers = []
            for cell in ws[header_row]:
                header = str(cell.value).strip() if cell.value else ''
                if header:
                    headers.append(header)
            
            logger.info(f"从Excel文件第{header_row}行提取到 {len(headers)} 个表头: {headers}")
            return headers
        
        except Exception as e:
            logger.error(f"提取Excel表头失败: {str(e)}")
            raise Exception(f"提取Excel表头失败: {str(e)}")
    
    def fuzzy_match_header(self, header: str) -> Optional[Tuple[str, str]]:
        """
        模糊匹配表头到系统字段
        返回: (system_key, label) 或 None
        """
        header_lower = header.lower().strip()
        
        # 精确匹配
        for system_key, field_info in SYSTEM_FIELD_MAPPINGS.items():
            if header == field_info['label']:
                logger.debug(f"精确匹配: '{header}' -> {system_key}")
                return (system_key, field_info['label'])
        
        # 模糊匹配：双向检查（表头包含关键词 或 关键词包含表头）
        best_match = None
        best_score = 0
        
        for system_key, field_info in SYSTEM_FIELD_MAPPINGS.items():
            for keyword in field_info['keywords']:
                keyword_lower = keyword.lower()
                
                # 双向匹配：
                # 1. 表头包含关键词（如"稿件编号"包含"稿件"）
                # 2. 关键词包含表头（如"稿件号"包含"稿件"）
                if keyword_lower in header_lower or header_lower in keyword_lower:
                    # 匹配度计算：
                    # - 如果完全匹配（表头=关键词），分数最高（1.0）
                    # - 否则根据共同部分的长度计算
                    if header_lower == keyword_lower:
                        score = 1.0
                    else:
                        # 计算共同部分的长度比例
                        common_length = min(len(header_lower), len(keyword_lower))
                        total_length = max(len(header_lower), len(keyword_lower))
                        score = common_length / total_length if total_length > 0 else 0
                    
                    if score > best_score:
                        best_score = score
                        best_match = (system_key, field_info['label'])
        
        if best_match:
            logger.debug(f"模糊匹配: '{header}' -> {best_match[0]} ({best_match[1]}, 匹配度: {best_score:.2f})")
            return best_match
        
        logger.debug(f"未匹配: '{header}'")
        return None
    
    def match_headers(self, headers: List[str]) -> List[Dict]:
        """
        批量匹配表头到系统字段
        返回: [{'template_header': '...', 'system_key': '...', 'label': '...', 'is_custom': False}, ...]
        """
        matched_headers = []
        
        for idx, header in enumerate(headers):
            match_result = self.fuzzy_match_header(header)
            
            if match_result:
                system_key, label = match_result
                matched_headers.append({
                    'template_header': header,
                    'system_key': system_key,
                    'label': label,
                    'order': idx + 1,
                    'is_custom': False
                })
            else:
                # 未匹配的作为自定义字段
                matched_headers.append({
                    'template_header': header,
                    'system_key': None,
                    'label': None,
                    'order': idx + 1,
                    'is_custom': True
                })
        
        return matched_headers
    
    def get_available_system_fields(self) -> List[Dict]:
        """
        获取所有可用的系统字段
        """
        return [
            {
                'key': key,
                'label': info['label'],
                'category': info.get('category', '其他'),
                'keywords': info['keywords']
            }
            for key, info in SYSTEM_FIELD_MAPPINGS.items()
        ]


